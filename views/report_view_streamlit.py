import streamlit as st
import pandas as pd
import io
import os
from openpyxl.utils import get_column_letter
from controllers.transaction_controller import TransactionController
from services.data_service import DataService 
from datetime import date  
from st_aggrid import AgGrid, GridOptionsBuilder, JsCode
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.drawing.image import Image

def export_to_excel(df, end_date):
    expected_cols = ["Nhóm", "Mã HH", "Tên hàng hóa", "Đvt", "Mức tối thiểu", "Tồn Đầu", "Nhập", "Xuất", "Tồn Cuối", "Ghi chú"]
    available_cols = [col for col in expected_cols if col in df.columns]
    df_export = df[available_cols].copy()

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        # 1. BẮT ĐẦU VIẾT BẢNG DỮ LIỆU TỪ DÒNG 7 (startrow=6 vì index từ 0)
        df_export.to_excel(writer, index=False, sheet_name='Data', startrow=6)
        ws = writer.sheets['Data']
        
        # Đóng băng dòng tiêu đề (Dòng 7 là tiêu đề, Dòng 8 là dữ liệu)
        ws.freeze_panes = 'A8' 
        
        # Bật bộ lọc (Filter) cho dữ liệu bắt đầu từ dòng 7
        ws.auto_filter.ref = f"A7:{get_column_letter(ws.max_column)}{ws.max_row}"
        
        # Tự động điều chỉnh độ rộng cột
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
            
        # =========================================================
        # [MỚI] TÔ MÀU XANH DƯƠNG CHO DÒNG TIÊU ĐỀ (DÒNG 7)
        # =========================================================
        # Định nghĩa màu nền xanh dương (Mã Hex: 0070C0) và chữ màu trắng in đậm
        header_fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=7, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # =========================================================
        # 2. THIẾT KẾ HEADER (THÔNG TIN CÔNG TY, LOGO, TIÊU ĐỀ)
        # =========================================================
        
        # Cột C1, C2, C3: Thông tin công ty
        ws['C1'] = "CÔNG TY TNHH THỦY SẢN PHÁT TIẾN"
        ws['C1'].font = Font(size=14,bold=True)
        ws['C2'] = "Địa chỉ: Lô B3, đường số 2, Cụm CN Mỹ Hiệp, Xã Mỹ Hiệp, Tỉnh Đồng Tháp"
        ws['C2'].font = Font(name="Arial", size=10)
        ws['C3'] = "Số điện thoại: 02778.553.388 - 02773.918.999"
        ws['C3'].font = Font(name="Arial", size=10)


        # Cột D4:E4: Merge Title
        ws.merge_cells('D4:E4')
        ws['D4'] = "BÁO CÁO TỒN KHO VẬT TƯ"
        ws['D4'].font = Font(bold=True, size=14)
        ws['D4'].alignment = Alignment(horizontal="center", vertical="center")

        # Cột D5:E5: Merge Ngày tháng (Động theo end_date)
        ws.merge_cells('D5:E5')
        if end_date:
            ws['D5'] = end_date.strftime("Ngày %d tháng %m năm %Y")
        ws['D5'].font = Font(italic=True)
        ws['D5'].alignment = Alignment(horizontal="center", vertical="center")

        # Cột A1: Chèn Logo (Cần có file logo.png nằm trong thư mục gốc của code)
        try:
            logo_path = "logo.png"
            if os.path.exists(logo_path):
                img = Image(logo_path)
                # Tùy chỉnh kích thước ảnh để vừa vặn (bạn có thể thay đổi số này)
                img.width = 100
                img.height = 100
                ws.add_image(img, 'A1')
        except Exception as e:
            print(f"Lỗi chèn ảnh Logo: {e}")

        # =========================================================
        # 3. TÔ MÀU ĐỎ CHO CÁC MẶT HÀNG CHẠM CẢNH BÁO TỒN KHO
        # =========================================================
        try:
            min_stock_col_idx = available_cols.index("Mức tối thiểu")
            end_stock_col_idx = available_cols.index("Tồn Cuối")
            
            red_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
            red_font = Font(color='D32F2F', bold=True)

            # Quét qua từng dòng dữ liệu (bắt đầu từ dòng 8)
            for row in ws.iter_rows(min_row=8, max_row=ws.max_row):
                min_val = row[min_stock_col_idx].value
                end_val = row[end_stock_col_idx].value
                
                try:
                    v_min = float(min_val) if min_val is not None else 0.0
                    v_end = float(end_val) if end_val is not None else 0.0
                    
                    if v_end <= v_min and v_min > 0:
                        for cell in row:
                            cell.fill = red_fill
                            cell.font = red_font
                except (ValueError, TypeError):
                    continue 
                    
        except ValueError:
            pass 

    return buffer.getvalue()

def export_history_to_excel(df):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='LichSu')
        ws = writer.sheets['LichSu']
        header_fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col)].width = 18
    return buffer.getvalue()

def show_report():
    st.subheader("Báo cáo tồn kho")
    
    DEFAULT_START_DATE = date(2026, 1, 1)
    
    if "clicked_report_filter" not in st.session_state:
        st.session_state.clicked_report_filter = False

    # =================================================================
    # --- [CẢI TIẾN CSS]: ÉP HÀNG NGANG VÀ KÉO SÁT TIÊU ĐỀ TRÊN ĐIỆN THOẠI ---
    # =================================================================
    st.markdown("""
        <style>
        /* 1. Kéo tiêu đề h3 dịch xuống hoặc thu nhỏ khoảng trống bên dưới nó */
        h3 {
            margin-bottom: -0.5rem !important;
            padding-bottom: 0rem !important;
        }
        
        /* 2. Ép 3 thành phần nằm ngang và kéo mạnh lên sát tiêu đề h3 */
        [data-testid="stHorizontalBlock"] {
            flex-wrap: nowrap !important;
            flex-direction: row !important;
            gap: 0.4rem !important; /* Thu hẹp khoảng cách giữa các ô */
            margin-top: -1.5rem !important; /* Lực hút đẩy dòng này lên sát tiêu đề */
        }
        
        /* Cho phép các ô tự động co nhỏ cho vừa màn hình điện thoại */
        [data-testid="stHorizontalBlock"] > div {
            min-width: 0px !important; 
        }
        
        /* Đẩy nút "Báo cáo" thụt xuống để thẳng hàng khít với ô nhập ngày */
        [data-testid="stHorizontalBlock"] > div:nth-child(3) {
            padding-top: 1.75rem !important; 
        }
        
        /* Làm nhỏ chữ tiêu đề "Từ ngày", "Đến ngày" để giao diện thanh thoát */
        [data-testid="stHorizontalBlock"] label {
            font-size: 0.8rem !important;
        }
        </style>
    """, unsafe_allow_html=True)

    col1, col2, col3 = st.columns([3, 3, 2.5])
    
    with col1:
        start_date = st.date_input("Từ ngày", value=DEFAULT_START_DATE)
    
    with col2:
        end_date = st.date_input("Đến ngày")
        
    with col3:
        if st.button("Báo cáo", type="primary", use_container_width=True):
            st.session_state.clicked_report_filter = True

    if st.session_state.clicked_report_filter:
        with st.spinner('Đang kết nối và xử lý dữ liệu...'):
            
            service = DataService()
            t_controller = TransactionController()
            
            products = service.get_products()
            df_h = t_controller.get_transaction_history().copy() 
            
            if not products or df_h is None or df_h.empty:
                st.warning("Không có dữ liệu giao dịch hoặc hàng hóa!")
                return

            # --- CHUẨN HÓA DỮ LIỆU ĐỂ TÍNH TOÁN (CA TỪ 06:00:00) ---
            try:
                # 1. Chuyển đổi sang datetime
                df_h['date'] = pd.to_datetime(df_h['Ngày'], dayfirst=True, format='mixed', errors='coerce')
                
                # 2. Tạo cột Ngày_Kho: Dịch lùi 6 tiếng để dồn mốc 06:00 về 00:00 của ngày hôm đó
                # Ví dụ: 02:00 sáng ngày 30/05 -> Trừ 6 tiếng -> 20:00 ngày 29/05 -> .normalize() -> 29/05/2026 00:00:00
                df_h['Ngày_Kho'] = (df_h['date'] - pd.Timedelta(hours=6)).dt.normalize()
                
                df_h['product_id'] = df_h['Mã HH'].astype(str).str.strip().str.upper()
                df_h['type'] = df_h['Loại'].astype(str).str.strip()
                df_h['qty'] = pd.to_numeric(df_h['Số Lượng'], errors='coerce').fillna(0)
            except KeyError as e:
                st.error(f"Lỗi cấu trúc cột trong Google Sheets: Thiếu cột {e}")
                return

            # 3. Chuẩn hóa ngày bắt đầu và kết thúc của người dùng
            start = pd.to_datetime(start_date).normalize()
            end = pd.to_datetime(end_date).normalize()
            
            # 4. Lọc dữ liệu dựa trên cột Ngày_Kho
            # Lũy kế từ đầu đến trước ngày bắt đầu (để tính tồn đầu)
            df_from_start = df_h[df_h['Ngày_Kho'] < start]
            if not df_from_start.empty:
                pivot_from_start = df_from_start.pivot_table(index='product_id', columns='type', values='qty', aggfunc='sum', fill_value=0)
                if 'Nhập' not in pivot_from_start.columns: pivot_from_start['Nhập'] = 0
                if 'Xuất' not in pivot_from_start.columns: pivot_from_start['Xuất'] = 0
            else:
                pivot_from_start = pd.DataFrame(columns=['Nhập', 'Xuất'])
                
            # Lọc dữ liệu trong khoảng thời gian được chọn
            df_period = df_h[(df_h['Ngày_Kho'] >= start) & (df_h['Ngày_Kho'] <= end)]
            if not df_period.empty:
                pivot_period = df_period.pivot_table(index='product_id', columns='type', values='qty', aggfunc='sum', fill_value=0)
                if 'Nhập' not in pivot_period.columns: pivot_period['Nhập'] = 0
                if 'Xuất' not in pivot_period.columns: pivot_period['Xuất'] = 0
            else:
                pivot_period = pd.DataFrame(columns=['Nhập', 'Xuất'])

            product_list = []
            for p in products:
                p_id = p[0]
                p_code = str(p[1]).strip()
                p_name = p[2]
                p_unit = p[3]
                p_stock = p[4]
                p_group = p[5] if len(p) > 5 else ""
                
                min_stock = float(p[6]) if len(p) > 6 and str(p[6]).strip() != "" else 0.0
                p_note = str(p[7]).strip() if len(p) > 7 else ""
                
                product_list.append([p_id, p_code, p_name, p_unit, p_stock, p_group, min_stock, p_note])
            
            df_products = pd.DataFrame(product_list, columns=["ID", "Mã HH", "Tên hàng hóa", "Đvt", "Tồn Hiện Tại", "Nhóm", "Mức tối thiểu", "Ghi chú"])
            df_products['Tồn Hiện Tại'] = pd.to_numeric(df_products['Tồn Hiện Tại'], errors='coerce').fillna(0)
            df_products['Mức tối thiểu'] = pd.to_numeric(df_products['Mức tối thiểu'], errors='coerce').fillna(0)
            df_products['Mã HH'] = df_products['Mã HH'].astype(str).str.strip().str.upper()
            
            df_report = df_products.merge(pivot_from_start[['Nhập', 'Xuất']], left_on='Mã HH', right_index=True, how='left').fillna(0)
            df_report.rename(columns={'Nhập': 'Nhập_Lũy_Kế', 'Xuất': 'Xuất_Lũy_Kế'}, inplace=True)
            
            df_report = df_report.merge(pivot_period[['Nhập', 'Xuất']], left_on='Mã HH', right_index=True, how='left').fillna(0)
            
            df_report['Tồn Đầu'] = df_report['Tồn Hiện Tại'] - df_report['Nhập_Lũy_Kế'] + df_report['Xuất_Lũy_Kế']
            df_report['Tồn Cuối'] = df_report['Tồn Đầu'] + df_report['Nhập'] - df_report['Xuất']
            
            df_report = df_report[["Nhóm", "Mã HH", "Tên hàng hóa", "Đvt", "Mức tối thiểu", "Tồn Đầu", "Nhập", "Xuất", "Tồn Cuối", "Ghi chú"]]
            
            st.markdown("---")
            st.subheader("🚨 Cảnh báo mức tồn kho")
            
            df_canh_bao = df_report[(df_report["Tồn Cuối"] <= df_report["Mức tối thiểu"]) & (df_report["Mức tối thiểu"] > 0)]
            
            if not df_canh_bao.empty:
                st.error(f"⚠️ Chú ý: Đang có **{len(df_canh_bao)}** mặt hàng chạm mức cảnh báo!")
                
                with st.expander("👇 Bấm vào đây để xem chi tiết danh sách hàng hóa", expanded=False):
                    st.dataframe(
                        df_canh_bao[["Mã HH", "Tên hàng hóa", "Đvt", "Tồn Cuối", "Mức tối thiểu"]], 
                        use_container_width=True, 
                        hide_index=True
                    )
            else:
                st.success("✅ Tuyệt vời! Tất cả hàng hóa đều có số lượng tồn kho trên mức an toàn.")
                
            st.markdown("---")
            st.subheader("📦 Chi tiết tồn kho toàn bộ hàng hóa")

            gb = GridOptionsBuilder.from_dataframe(df_report)
            gb.configure_default_column(sortable=True, filter=True, resizable=True, flex=1, minWidth=100)
            
            gb.configure_column("Nhóm", rowGroup=True, hide=True)
            
            gb.configure_column("Mã HH", minWidth=60, maxWidth=120, cellStyle={'textAlign': 'center'})
            gb.configure_column("Tên hàng hóa", minWidth=150, cellStyle={'textAlign': 'left'})
            gb.configure_column("Đvt", minWidth=60, maxWidth=100, cellStyle={'textAlign': 'center'})
            
            gb.configure_column("Ghi chú", minWidth=120, cellStyle={'textAlign': 'left'})

            for col_name in ["Mức tối thiểu", "Tồn Đầu", "Nhập", "Xuất", "Tồn Cuối"]:
                gb.configure_column(
                    col_name,
                    minWidth=90, maxWidth=130,
                    type=["numericColumn"],
                    filter='agNumberColumnFilter',
                    valueFormatter="Number(x).toLocaleString('en-US')",
                    cellStyle={'textAlign': 'right'})
            
            row_style_jscode = JsCode("""
            function(params) {
                if (params.data && params.data['Tồn Cuối'] <= params.data['Mức tối thiểu'] && params.data['Mức tối thiểu'] > 0) {
                    return {
                        'backgroundColor': '#ffe6e6', 
                        'color': '#d32f2f',           
                        'fontWeight': 'bold'
                    };
                }
                return null;
            }
            """)
            gb.configure_grid_options(getRowStyle=row_style_jscode)
            
            go = gb.build() 
            
            AgGrid(
                df_report,
                gridOptions=go,
                fit_columns_on_grid_load=True,
                theme='streamlit',
                allow_unsafe_jscode=True, 
                height=650)
            
            st.markdown("<br>", unsafe_allow_html=True)
            st.download_button(
                label="📥 Xuất báo cáo ra Excel (.xlsx)",
                data=export_to_excel(df_report, end_date), 
                file_name=f"BaoCaoTonKho_{end_date.strftime('%d%m%Y')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            
def get_count_import_by_department(service, start_date, end_date):
    """Tính số lần nhập kho theo Bộ phận (cột I trong Transactions)"""
    history_df = service.get_history()
    if history_df is None or history_df.empty:
        return pd.DataFrame()

    df = history_df.copy()
    # Chuyển Ngày thành datetime
    df['date_obj'] = pd.to_datetime(df['Ngày'], dayfirst=True, errors='coerce')
    # Áp dụng quy tắc mốc 06:00 sáng
    df['Ngày_Kho'] = (df['date_obj'] - pd.Timedelta(hours=6)).dt.date
    
    # 1. Lọc chỉ lấy giao dịch NHẬP
    df_nhap = df[df['Loại'].astype(str).str.strip().str.upper() == 'NHẬP'].copy()
    
    # 2. Lọc theo khoảng thời gian
    start = pd.to_datetime(start_date).date()
    end = pd.to_datetime(end_date).date()
    df_nhap = df_nhap[(df_nhap['Ngày_Kho'] >= start) & (df_nhap['Ngày_Kho'] <= end)]
    
    # 3. Sử dụng cột 'Bộ phận' để thống kê
    # Đảm bảo tên cột khớp với file Google Sheets của bạn (ví dụ: 'Bộ phận')
    col_bp = 'Bộ phận' 
    
    if col_bp not in df_nhap.columns:
        # Nếu cột không tồn tại, trả về thông báo lỗi
        return pd.DataFrame(columns=[col_bp, 'Số lần nhập'])

    # Điền giá trị rỗng nếu cột 'Bộ phận' có ô trống
    df_nhap[col_bp] = df_nhap[col_bp].fillna("Chưa xác định")
    
    # 4. Loại bỏ trùng lặp: Cùng ngày và cùng bộ phận chỉ tính 1 lần
    df_unique = df_nhap.drop_duplicates(subset=['Ngày_Kho', col_bp])
    
    # 5. Đếm số lần nhập theo bộ phận
    report = df_unique.groupby(col_bp).size().reset_index(name='Số lần nhập')
    return report