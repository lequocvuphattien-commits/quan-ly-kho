import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as OpenpyxlImage
import os
from io import BytesIO
import datetime
import io
from openpyxl.utils import get_column_letter

# --- 1. HÀM TẠO FILE EXCEL NGẦM ---
def export_phieu_xuat_excel(export_data, selected_date, department_name):
    """
    Hàm xuất dữ liệu giỏ hàng ra file Excel - CÓ LOGO CÔNG TY & TỐI ƯU IN KHỔ A4 DỌC
    """
    template_path = "phieu_mau.xlsx"
    try:
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active # BỔ SUNG: Khai báo ws khi mở file mẫu thành công
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
    
    # --- CẤU HÌNH TRANG IN (VỪA KHÍT A4 DỌC) ---
    ws.views.sheetView[0].showGridLines = False 
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1 
    ws.page_setup.fitToHeight = 0 
    
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.6
    ws.page_margins.bottom = 0.6

    # --- ĐỊNH NGHĨA FONT & STYLE ---
    font_regular = Font(name="Arial", size=11)
    font_bold = Font(name="Arial", size=11, bold=True)
    font_italic = Font(name="Arial", size=11, italic=True)
    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid") 
    
    thin_border = Border(
        left=Side(style='thin', color='B0B0B0'),
        right=Side(style='thin', color='B0B0B0'),
        top=Side(style='thin', color='B0B0B0'),
        bottom=Side(style='thin', color='B0B0B0')
    )
    
    # --- XỬ LÝ CHÈN LOGO ---
    logo_path = "logo.png"
    if os.path.exists(logo_path):
        img = OpenpyxlImage(logo_path)
        img.width = 125
        img.height = 75
        ws.add_image(img, 'A4')
        
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18
    
    ws['A1'] = "CÔNG TY TNHH THỦY SẢN PHÁT TIẾN"
    ws['A1'].font = Font(name="Arial", size=11, bold=True)
    ws['A2'] = "Địa chỉ: Lô B3, đường số 2, Cụm CN Mỹ Hiệp, Xã Mỹ Hiệp, Tỉnh Đồng Tháp"
    ws['A2'].font = Font(name="Arial", size=10, italic=True)
    ws['A3'] = "Số điện thoại: 02778.553.388 - 02773.918.999"
    ws['A3'].font = Font(name="Arial", size=10, italic=True)
    ws['E7'] = "Số phiếu:....................."
    ws['E7'].font = Font(name="Arial", size=11, bold=True)

    # --- TIÊU ĐỀ PHIẾU (DÒNG 5) ---
    ws['C5'] = "PHIẾU XUẤT KHO"
    ws['C5'].font = Font(name="Arial", size=16, bold=True, color="1F4E78")
    ws['C5'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 32
    
    # --- BỘ PHẬN ĐỀ NGHỊ (Linh hoạt theo Selectbox) ---
    ws['A7'] = f"Bộ phận đề nghị: {department_name}"
    ws['A7'].font = font_bold
    ws.row_dimensions[7].height = 20

    ws['A7'] = f"Bộ phận đề nghị: {department_name}"
    ws['A7'].font = font_bold
    ws.row_dimensions[7].height = 20
       
    # --- HEADER BẢNG DỮ LIỆU ---
    ws.row_dimensions[9].height = 26
    
    # ĐÃ THÊM CỘT GHI CHÚ
    headers = ["STT", "Tên hàng hóa", "Đvt", "Số Lượng", "Diễn Giải", "Ghi chú"]
    cols = ["A", "B", "C", "D", "E", "F"]
    
    for col_letter, header_text in zip(cols, headers):
        cell = ws[f"{col_letter}9"]
        cell.value = header_text
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    # --- ĐỔ DỮ LIỆU BẢNG ---
    start_row = 10
    current_row = start_row
    
    for i, item in enumerate(export_data):
        current_row = start_row + i
        ws.row_dimensions[current_row].height = 20
        
        ws[f"A{current_row}"] = i + 1
        ws[f"B{current_row}"] = item.get("Tên HH", "")
        ws[f"C{current_row}"] = item.get("Đvt", "")
        ws[f"D{current_row}"] = float(item.get("Số lượng", 0))
        ws[f"E{current_row}"] = str(item.get("Diễn Giải", ""))
        ws[f"F{current_row}"] = "" # Để trống hoàn toàn cột F để ghi tay
                
        # Cấu hình căn lề
        ws[f"A{current_row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"B{current_row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"C{current_row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"D{current_row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"D{current_row}"].number_format = '#,##0' 
        ws[f"E{current_row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"F{current_row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
              
        for col_letter in cols:
            cell = ws[f"{col_letter}{current_row}"]
            cell.font = font_regular
            cell.border = thin_border
            if i % 2 == 1:
                cell.fill = fill_zebra
                
    # --- CHỮ KÝ & NGÀY THÁNG ---
    last_data_row = current_row
    date_row = last_data_row + 2
    sign_title_row = date_row + 1
    
    # Ghi ngày tháng (Dịch sang cột E và F để cân đối với 6 cột)
    date_str = f"Ngày {selected_date.day:02d} tháng {selected_date.month:02d} năm {selected_date.year}"
    ws.merge_cells(f'E{date_row}:F{date_row}')
    ws[f'E{date_row}'] = date_str
    ws[f'E{date_row}'].font = font_italic
    ws[f'E{date_row}'].alignment = Alignment(horizontal="center", vertical="center")
    
    ws.row_dimensions[sign_title_row].height = 25
    
    # Chữ ký Kế Toán
    ws[f"B{sign_title_row}"] = "Kế Toán"
    ws[f"B{sign_title_row}"].font = font_bold
    ws[f"B{sign_title_row}"].alignment = Alignment(horizontal="left", vertical="center")
    
    # Chữ ký Thủ Kho
    ws.merge_cells(f'C{sign_title_row}:D{sign_title_row}')
    ws[f"C{sign_title_row}"] = "Thủ Kho"
    ws[f"C{sign_title_row}"].font = font_bold
    ws[f"C{sign_title_row}"].alignment = Alignment(horizontal="left", vertical="center")
    
    # Chữ ký Người lập (Dịch sang cột E và F)
    ws.merge_cells(f'E{sign_title_row}:F{sign_title_row}')
    ws[f'E{sign_title_row}'] = "   Người Lập"
    ws[f'E{sign_title_row}'].font = font_bold
    ws[f'E{sign_title_row}'].alignment = Alignment(horizontal="center", vertical="center")
        
    # --- CHỈNH KÍCH THƯỚC CỘT (Tối ưu lại tỷ lệ cho 6 cột vừa khổ A4) ---
    ws.column_dimensions['A'].width = 6   # STT
    ws.column_dimensions['B'].width = 38  # Tên hàng hóa
    ws.column_dimensions['C'].width = 8   # Đvt
    ws.column_dimensions['D'].width = 11  # Số lượng
    ws.column_dimensions['E'].width = 18  # Diễn giải
    ws.column_dimensions['F'].width = 15  # Ghi chú
        
    output = BytesIO()
    wb.save(output)
    return output.getvalue()

# --- 2. HÀM HIỂN THỊ GIAO DIỆN (VIEW) ---
def show_print_export_view(service):
    st.subheader("🖨️ In Phiếu Xuất Kho")
    
    col_date, col_dept = st.columns(2)
    with col_date:
        selected_date = st.date_input("📅 Chọn ngày in phiếu:", datetime.date.today())
    
    # 1. Lấy và chuẩn hóa dữ liệu với mốc 6h sáng
    history_data = service.get_history()
    if history_data is None or history_data.empty:
        st.warning("Không có dữ liệu lịch sử giao dịch!")
        return
        
    df = history_data.copy()
    # Chuyển Ngày thành datetime
    df['date_obj'] = pd.to_datetime(df['Ngày'], dayfirst=True, errors='coerce')
    # Tạo cột Ngày_Kho (dịch lùi 6 tiếng để dồn về mốc 06:00)
    df['Ngày_Kho'] = (df['date_obj'] - pd.Timedelta(hours=6)).dt.date
    
    # Lọc các giao dịch XUẤT theo Ngày_Kho đã chọn
    filtered_df = df[(df['Loại'].astype(str).str.strip().str.upper() == 'XUẤT') & 
                     (df['Ngày_Kho'] == selected_date)]
    
    if filtered_df.empty:
        st.warning(f"⚠️ Không có giao dịch XUẤT KHO nào trong ngày {selected_date.strftime('%d/%m/%Y')} (tính từ 06:00).")
        return

    # 2. Xử lý Bộ phận đề nghị (Selectbox)
    dien_giai_list = filtered_df["Diễn Giải"].astype(str).str.strip().unique().tolist()
    clean_dg_list = ['Không có diễn giải' if dg.lower() in ['nan', ''] else dg for dg in dien_giai_list]
    clean_dg_list = sorted(list(set(clean_dg_list)))
    
    with col_dept:
        department_name = st.selectbox("🏢 Chọn bộ phận đề nghị:", clean_dg_list)
        
    # Lọc lại lần 2 theo Bộ phận
    if department_name == 'Không có diễn giải':
        final_df = filtered_df[filtered_df["Diễn Giải"].astype(str).str.strip().isin(['', 'nan', 'NaN'])]
    else:
        final_df = filtered_df[filtered_df["Diễn Giải"].astype(str).str.strip() == department_name]

    # 3. Gộp dòng và chuẩn bị dữ liệu xuất
    products = service.get_products()
    dvt_dict = {str(p[1]): str(p[3]) for p in products} if products else {}
    
    raw_data = []
    for _, row in final_df.iterrows():
        ma_hh = str(row.get("Mã HH", ""))
        raw_data.append({
            "Tên HH": row.get("Tên hàng hóa", ma_hh),
            "Đvt": dvt_dict.get(ma_hh, ""),
            "Số lượng": float(row.get("Số Lượng", 0)),
            "Diễn Giải": str(row.get("Diễn Giải", ""))
        })
    
    df_grouped = pd.DataFrame(raw_data).groupby(['Tên HH', 'Đvt', 'Diễn Giải'], as_index=False)['Số lượng'].sum()
    df_grouped.insert(0, 'Chọn', True)
    
    # 4. Giao diện tích chọn
    edited_df = st.data_editor(
        df_grouped, use_container_width=True, hide_index=True,
        disabled=["Tên HH", "Đvt", "Số lượng", "Diễn Giải"],
        column_config={"Số lượng": st.column_config.NumberColumn("Số Lượng", format="%d")}
    )
    
    selected_df = edited_df[edited_df["Chọn"] == True]
    if selected_df.empty:
        st.error("🚫 Bạn chưa chọn mặt hàng nào!")
        return
        
    # 5. Xuất Excel
    export_data = selected_df.drop(columns=['Chọn']).to_dict('records')
    excel_data = export_phieu_xuat_excel(export_data, selected_date, department_name)
    
    st.download_button(
        label="📥 TẢI FILE EXCEL PHIẾU XUẤT",
        data=excel_data,
        file_name=f"Phieu_Xuat_{selected_date.strftime('%d%m%Y')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary", use_container_width=True)
    
# ĐẢM BẢO HÀM NÀY NẰM NGOÀI CÙNG (DÁN ĐÈ ĐOẠN NÀY VÀO CUỐI FILE)
def export_history_to_excel(df):
    """Xuất lịch sử giao dịch ra file Excel"""
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='LichSu')
        ws = writer.sheets['LichSu']
        
        # Định dạng tiêu đề
        header_fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col)].width = 18
            
    return buffer.getvalue()