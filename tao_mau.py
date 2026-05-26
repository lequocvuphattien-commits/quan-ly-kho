import openpyxl
from openpyxl.styles import Font, Alignment

def tao_file_mau_chuan():
    # Tạo một file Excel mới
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mau_Phieu_Xuat"
    
    # Tắt đường kẻ lưới mặc định của Excel cho đẹp
    ws.views.sheetView[0].showGridLines = False

    # 1. Tên công ty và địa chỉ
    ws['A1'] = "CÔNG TY TNHH QUẢN LÝ KHO ĐÔNG THÁP"
    ws['A1'].font = Font(name="Arial", size=12, bold=True)
    
    ws['A2'] = "Địa chỉ: Cao Lãnh, Đồng Tháp, Việt Nam"
    ws['A2'].font = Font(name="Arial", size=10, italic=True)

    # 2. Tiêu đề phiếu
    ws.merge_cells('A4:F4') # Gộp ô từ A4 đến F4
    ws['A4'] = "PHIẾU XUẤT KHO"
    ws['A4'].font = Font(name="Arial", size=16, bold=True)
    ws['A4'].alignment = Alignment(horizontal="center", vertical="center")

    # Lưu file
    wb.save("phieu_mau.xlsx")
    print("✅ Đã tạo thành công file 'phieu_mau.xlsx' chuẩn!")

# Chạy hàm
tao_file_mau_chuan()