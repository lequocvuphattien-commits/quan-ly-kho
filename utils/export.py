import pandas as pd
import io
from openpyxl.utils import get_column_letter

def export_to_excel(df):
    """
    Hàm hỗ trợ xuất DataFrame ra file Excel (định dạng .xlsx).
    """
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Data')
        
        # Định dạng thêm cho đẹp
        worksheet = writer.sheets['Data']
        worksheet.freeze_panes = 'A2'
        
        # Tự động chỉnh độ rộng cột
        for col in range(1, worksheet.max_column + 1):
            column_letter = get_column_letter(col)
            worksheet.column_dimensions[column_letter].width = 15
            
    return buffer.getvalue()